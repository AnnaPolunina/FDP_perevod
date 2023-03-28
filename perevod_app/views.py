from openpyxl import load_workbook
import json
from django.shortcuts import render
from .models import Translation
from django.views import View
from django.views.generic.list import ListView
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from .models import Translation


class WordsBaseView(View):
    '''базовый класс'''
    model = Translation
    fields = '__all__'
    success_url = reverse_lazy('all')

class WordsListView(WordsBaseView, ListView):
    """Список всех слов"""

class WordsCreateView(WordsBaseView, CreateView):
    """Создание новой пары слов"""

class WordsUpdateView(WordsBaseView, UpdateView):
    """Редактирование пары слов"""

class WordsDeleteView(WordsBaseView, DeleteView):
    """Удаление пары слов"""

def search_word(request):
    user_value  = request.GET.get('search_word')
    add_words_from_exel()
    add_words_from_DB_in_exel()
    json_from_bd = json_value()
    print(json_from_bd)
    try:
        result_BD = Translation.objects.get(en=user_value)
        if result_BD:
            result = '{0}-{1}'.format(result_BD.en, result_BD.uk)
            return render(request, 'perevod_app/search_word.html', {'result': result, 'json_from_bd':json_from_bd})
        else:
            result = 'Слово не знайдено'
        return render(request, 'perevod_app/search_word.html', {'result': result, 'json_from_bd':json_from_bd})

    except:
        result = 'Слово не знайдено'
        return render(request, 'perevod_app/search_word.html', {'result': result, 'json_from_bd':json_from_bd})

def add_words_from_exel():
    '''сохранение всех записей с exel в БД'''
    wb = load_workbook(filename='translations.xlsx')
    sheet = wb['Sheet1']
    for row in sheet.iter_rows(values_only=True):
        try:
            trans_bd = Translation()
            trans_bd.en=row[0]
            trans_bd.uk=row[1]
            trans_bd.save()
        except:
            continue

def add_words_from_DB_in_exel():
    '''сохранение всех записей с БД в еxel'''
    wb = load_workbook(filename='translations.xlsx')
    sheet = wb['Sheet1']
    trans_bd = Translation.objects.all()
    exel_list =[]
    for row in sheet.iter_rows(values_only=True):
        exel_list.append(row[0])
    for word in trans_bd:
        if word.en not in exel_list:
            sheet.append({'A':word.en, 'B':word.uk})
            wb.save('translations.xlsx')

'''def json_value():
    bd_value = Translation.objects.all()
    main_dict = {}
    for value in bd_value:
        main_dict[value.en] = value.uk
    return json.dumps(main_dict)'''

def json_value():
    bd_value = Translation.objects.all()
    main_list = []
    for value in bd_value:
        main_list.append(value.en)    
    return main_list




